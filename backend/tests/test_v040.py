"""V0.4.0 LAN mode, identity, AI fallback and backup regression tests.

Every test is self-contained: no real Qwen request is made, process-level
configuration is probed in child interpreters, and all databases/files live
under temporary directories.
"""
from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path
import sqlite3
import subprocess
import sys
import tempfile
from types import SimpleNamespace
import unittest
from unittest import mock

from fastapi import FastAPI, HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool


PROJECT_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = PROJECT_ROOT / "backend"
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(BACKEND_ROOT))

from app import config
from app.api import handovers, session as session_api
from app.db import Base, get_db
from app.migrations import migrate_database
from app.models import (
    AuditEvent,
    ExternalAssessment,
    HandoverBatch,
    HandoverItem,
    HandoverStationMeta,
    ImportJob,
    SectionImportPreview,
    Staff,
    Station,
)
from app.services import handover_service
from app.services import backup as backup_service
from app.services.ai.adapter import MockAdapter, QwenAdapter, ResilientAdapter
from app.services.importer import sections
import server_config
from server_migration import migrate_v030_data, relocate_server_data


class V040ConfigurationTest(unittest.TestCase):
    """Probe import-time configuration without reloading shared modules."""

    def _probe(self, mode: str, public_host: str = "") -> dict:
        script = r"""
import json
from sqlalchemy import text
from app import config
from app.db import engine
with engine.connect() as connection:
    pragmas = {
        "foreign_keys": int(connection.execute(text("PRAGMA foreign_keys")).scalar()),
        "busy_timeout": int(connection.execute(text("PRAGMA busy_timeout")).scalar()),
        "journal_mode": str(connection.execute(text("PRAGMA journal_mode")).scalar()).lower(),
        "synchronous": int(connection.execute(text("PRAGMA synchronous")).scalar()),
    }
print(json.dumps({
    "mode": config.APP_MODE,
    "host": config.APP_HOST,
    "port": config.APP_PORT,
    "public_url": config.PUBLIC_URL,
    "auth_required": config.AUTH_REQUIRED,
    "database": config.DATABASE_URL,
    "pragmas": pragmas,
}))
"""
        with tempfile.TemporaryDirectory() as tmp:
            env = os.environ.copy()
            for name in list(env):
                if name.startswith(("JX_", "AI_", "QWEN_")) or name in {
                    "WORD_TEMPLATE", "CLOUD_PUBLISH_DIR",
                }:
                    env.pop(name, None)
            env.update({
                "PYTHONPATH": str(BACKEND_ROOT),
                "JX_HANDOVER_MODE": mode,
                "JX_HANDOVER_DATA_DIR": tmp,
                # Invalid optional values must not make the service unbootable.
                "AI_TIMEOUT_SECONDS": "not-a-number",
                "JX_SESSION_TTL_HOURS": "not-a-number",
            })
            if public_host:
                env["JX_PUBLIC_HOST"] = public_host
            completed = subprocess.run(
                [sys.executable, "-c", script],
                cwd=PROJECT_ROOT,
                env=env,
                check=True,
                capture_output=True,
                text=True,
                timeout=30,
            )
        return json.loads(completed.stdout.strip().splitlines()[-1])

    def test_desktop_and_server_modes_keep_fixed_port_and_isolated_env(self):
        before = {
            key: os.environ.get(key)
            for key in ("JX_HANDOVER_MODE", "JX_PUBLIC_HOST", "JX_HANDOVER_DATA_DIR")
        }

        desktop = self._probe("desktop")
        server = self._probe("server", "192.168.14.88")

        self.assertEqual(desktop["host"], "127.0.0.1")
        self.assertEqual(desktop["port"], 8765)
        self.assertEqual(desktop["public_url"], "http://127.0.0.1:8765")
        self.assertFalse(desktop["auth_required"])
        self.assertEqual(server["host"], "0.0.0.0")
        self.assertEqual(server["port"], 8765)
        self.assertEqual(server["public_url"], "http://192.168.14.88:8765")
        self.assertTrue(server["auth_required"])
        for payload in (desktop, server):
            self.assertEqual(payload["pragmas"]["journal_mode"], "wal")
            self.assertEqual(payload["pragmas"]["foreign_keys"], 1)
            self.assertGreaterEqual(payload["pragmas"]["busy_timeout"], 30000)
            # SQLite reports NORMAL as integer 1.
            self.assertEqual(payload["pragmas"]["synchronous"], 1)

        self.assertEqual(
            before,
            {
                key: os.environ.get(key)
                for key in before
            },
        )

    def test_server_live_data_root_is_customizable_but_network_paths_are_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            custom_root = Path(tmp) / "自定义正式数据"
            validated = server_config.validate_local_data_root(custom_root)
            self.assertEqual(validated, custom_root.resolve())
            self.assertTrue(validated.is_dir())
            self.assertFalse(any(validated.glob(".jxhandover-write-test-*.tmp")))

            settings = dict(server_config.DEFAULT_SETTINGS)
            settings["data_root"] = str(custom_root)
            secrets_value = dict(server_config.DEFAULT_SECRETS)
            with mock.patch.dict(os.environ, {}, clear=False):
                server_config.apply_server_environment(settings, secrets_value)
                self.assertEqual(
                    Path(os.environ["JX_HANDOVER_DATA_DIR"]), custom_root.resolve()
                )

        with self.assertRaisesRegex(ValueError, "共享目录"):
            server_config.validate_local_data_root(
                r"\\192.168.14.52\江西片区检修中心\交接班系统",
                create=False,
            )
        with mock.patch.object(server_config, "_windows_drive_type", return_value=4):
            with self.assertRaisesRegex(ValueError, "映射网络盘"):
                server_config.validate_local_data_root(
                    r"Z:\交接班系统",
                    create=False,
                )


class V040MigrationTest(unittest.TestCase):
    def test_v040_migration_adds_ai_columns_and_audit_table_idempotently(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            database = root / "v030.db"
            engine = create_engine(f"sqlite:///{database.as_posix()}")
            Base.metadata.create_all(engine)
            db = sessionmaker(bind=engine)()
            station = Station(code="MIG", name="迁移测试场站", aliases_json="[]")
            db.add(station)
            db.flush()
            batch = HandoverBatch(
                start_date="2026-08-01",
                end_date="2026-08-10",
                handover_date="2026-08-10",
            )
            db.add(batch)
            db.flush()
            meta = HandoverStationMeta(batch_id=batch.id, station_id=station.id)
            job = ImportJob(source_type="section_xlsx", file_name="old.xlsx")
            db.add_all([meta, job])
            db.flush()
            preview = SectionImportPreview(
                batch_id=batch.id,
                station_meta_id=meta.id,
                import_job_id=job.id,
                parser_key="work_log",
                source_file_name="old.xlsx",
                source_sha256="abc",
                normalized_json='[{"legacy":true}]',
            )
            db.add(preview)
            db.commit()
            preview_id = preview.id
            db.close()

            with engine.begin() as connection:
                connection.execute(text("DROP TABLE audit_events"))
                connection.execute(text(
                    "ALTER TABLE section_import_previews DROP COLUMN ai_status"
                ))
                connection.execute(text(
                    "ALTER TABLE section_import_previews DROP COLUMN ai_model"
                ))
                connection.execute(text(
                    "ALTER TABLE section_import_previews DROP COLUMN ai_usage_json"
                ))

            backup_dir = root / "migration-backups"
            first = migrate_database(engine, backup_dir=backup_dir)
            inspector = inspect(engine)
            columns = {
                column["name"]
                for column in inspector.get_columns("section_import_previews")
            }
            self.assertTrue(
                {"ai_status", "ai_model", "ai_usage_json"}.issubset(columns)
            )
            self.assertIn("audit_events", inspector.get_table_names())
            self.assertIsNotNone(first["backup_path"])
            self.assertTrue(Path(first["backup_path"]).exists())
            with engine.connect() as connection:
                migrated = connection.execute(text(
                    "SELECT source_file_name, normalized_json, ai_status, "
                    "ai_model, ai_usage_json FROM section_import_previews "
                    "WHERE id=:id"
                ), {"id": preview_id}).one()
            self.assertEqual(migrated.source_file_name, "old.xlsx")
            self.assertEqual(migrated.normalized_json, '[{"legacy":true}]')
            self.assertEqual(migrated.ai_status, "not_requested")
            self.assertEqual(migrated.ai_model, "")
            self.assertEqual(migrated.ai_usage_json, "{}")

            backup_count = len(list(backup_dir.glob("*.db")))
            second = migrate_database(engine, backup_dir=backup_dir)
            self.assertEqual(second["changed"], [])
            self.assertIsNone(second["backup_path"])
            self.assertEqual(len(list(backup_dir.glob("*.db"))), backup_count)
            engine.dispose()


class V040SessionTest(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine)
        db = self.Session()
        db.add_all([
            Staff(
                station_code="REGION",
                name="钟宇",
                role="科技专责",
                note="内部岗位信息",
            ),
            Staff(
                station_code="REGION",
                name="熊思奇",
                role="带班负责人",
                note="",
            ),
        ])
        db.commit()
        db.close()

        self.app = FastAPI()
        self.app.include_router(session_api.router)
        self.app.include_router(handovers.router)

        def override_db():
            db = self.Session()
            try:
                yield db
            finally:
                db.close()

        self.app.dependency_overrides[get_db] = override_db

    def tearDown(self) -> None:
        self.engine.dispose()

    def test_login_options_show_names_only_and_protected_routes_require_cookie(self):
        settings = {
            "AUTH_REQUIRED": True,
            "ACCESS_CODE": "lan-only-code",
            "SESSION_SECRET": "fixed-test-secret-that-never-leaves-this-test",
            "SESSION_TTL_HOURS": 1,
            "ADMIN_NAMES": {"钟宇"},
            "APP_MODE": "server",
        }
        with mock.patch.multiple(config, **settings):
            with TestClient(self.app) as anonymous:
                protected = anonymous.get("/api/handovers")
                self.assertEqual(protected.status_code, 401)

                options = anonymous.get("/api/session/options")
                self.assertEqual(options.status_code, 200)
                self.assertEqual(options.json()["staff_names"], ["熊思奇", "钟宇"])
                serialized = json.dumps(options.json(), ensure_ascii=False)
                self.assertNotIn("科技专责", serialized)
                self.assertNotIn("带班负责人", serialized)
                self.assertNotIn("（", serialized)

                denied = anonymous.post("/api/session/login", json={
                    "name": "钟宇", "access_code": "wrong",
                })
                self.assertEqual(denied.status_code, 401)

            with TestClient(self.app) as authenticated:
                login = authenticated.post("/api/session/login", json={
                    "name": "钟宇", "access_code": "lan-only-code",
                })
                self.assertEqual(login.status_code, 200)
                self.assertEqual(login.json()["name"], "钟宇")
                self.assertEqual(login.json()["role"], "admin")
                self.assertIn("httponly", login.headers["set-cookie"].lower())

                self.assertEqual(authenticated.get("/api/handovers").status_code, 200)
                staff = authenticated.get("/api/staff")
                self.assertEqual(staff.status_code, 200)
                # The UI picker hides roles, while the staff API still retains them.
                self.assertEqual(staff.json()[0]["role"], "科技专责")

                logout = authenticated.post("/api/session/logout")
                self.assertEqual(logout.status_code, 200)
                self.assertEqual(authenticated.get("/api/handovers").status_code, 401)


class V040OptimisticLockTest(unittest.TestCase):
    def test_station_meta_rejects_stale_revision_with_409(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        db = sessionmaker(bind=engine)()
        station = Station(code="LOCK", name="并发测试场站", aliases_json="[]")
        db.add(station)
        db.flush()
        batch = HandoverBatch(
            start_date="2026-08-14",
            end_date="2026-08-23",
            handover_date="2026-08-23",
        )
        db.add(batch)
        db.flush()
        meta = HandoverStationMeta(batch_id=batch.id, station_id=station.id)
        db.add(meta)
        db.commit()

        first = handover_service.patch_station_meta(
            db, meta.id, 1, {"duty_leader": "甲"}
        )
        self.assertEqual(first["revision"], 2)
        with self.assertRaises(HTTPException) as conflict:
            handover_service.patch_station_meta(
                db, meta.id, 1, {"duty_leader": "乙"}
            )
        self.assertEqual(conflict.exception.status_code, 409)
        self.assertEqual(conflict.exception.detail["code"], "REVISION_CONFLICT")
        self.assertEqual(conflict.exception.detail["current_revision"], 2)
        db.close()
        engine.dispose()


class V040QwenAdapterTest(unittest.TestCase):
    @staticmethod
    def _response(content: dict):
        return SimpleNamespace(
            choices=[SimpleNamespace(
                message=SimpleNamespace(content=json.dumps(content, ensure_ascii=False))
            )],
            usage=SimpleNamespace(
                prompt_tokens=12,
                completion_tokens=7,
                total_tokens=19,
            ),
        )

    def test_qwen_uses_strict_json_schema_and_non_thinking_mode(self):
        fake_client = mock.Mock()
        fake_client.chat.completions.create.return_value = self._response({
            "items": [{
                "preview_key": "row-1",
                "title_snapshot": "整理后的标题",
                "section": "important",
                "status": "completed",
                "priority": "important",
                "completed_by": "甲",
                "previous_owner": "",
                "next_owner": "",
                "summary": "按方案执行",
                "latest_progress": "已完成",
                "blocker": "",
                "next_action": "",
                "confidence": 0.92,
                "warnings": [],
            }],
        })
        with mock.patch.multiple(
            config,
            QWEN_API_KEY="fake-key",
            QWEN_BASE_URL="https://example.invalid/v1",
            QWEN_MODEL="qwen3.8-flash",
            AI_TIMEOUT_SECONDS=9.0,
        ), mock.patch("openai.OpenAI", return_value=fake_client) as client_factory:
            adapter = QwenAdapter()
            result = adapter.enrich_preview_rows([
                {
                    "preview_key": "row-1",
                    "kind": "item",
                    "title_snapshot": "原始标题",
                    "section": "handover",
                    "status": "in_progress",
                    "priority": "normal",
                    "start_date": "2026-08-14",
                    "end_date": "2026-08-20",
                },
                {
                    "preview_key": "external-1",
                    "kind": "external",
                    "contractor": "不应发送给此模型流程",
                },
            ], {"station": "测试场站"})

        client_factory.assert_called_once_with(
            api_key="fake-key",
            base_url="https://example.invalid/v1",
            timeout=9.0,
        )
        kwargs = fake_client.chat.completions.create.call_args.kwargs
        self.assertEqual(kwargs["model"], "qwen3.8-flash")
        self.assertEqual(kwargs["extra_body"], {"enable_thinking": False})
        self.assertEqual(kwargs["response_format"]["type"], "json_schema")
        schema_envelope = kwargs["response_format"]["json_schema"]
        self.assertTrue(schema_envelope["strict"])
        self.assertFalse(schema_envelope["schema"]["additionalProperties"])
        sent_payload = json.loads(kwargs["messages"][1]["content"])
        self.assertEqual(len(sent_payload["rows"]), 1)
        self.assertEqual(sent_payload["rows"][0]["preview_key"], "row-1")
        self.assertEqual(result["items"][0]["title_snapshot"], "整理后的标题")
        self.assertEqual(result["usage"]["total_tokens"], 19)

    def test_qwen_exception_falls_back_without_network_or_workflow_failure(self):
        fake_client = mock.Mock()
        fake_client.chat.completions.create.side_effect = RuntimeError("mock outage")
        with mock.patch.multiple(
            config,
            QWEN_API_KEY="fake-key",
            QWEN_BASE_URL="https://example.invalid/v1",
            QWEN_MODEL="qwen3.8-flash",
        ), mock.patch(
            "openai.OpenAI", return_value=fake_client
        ), mock.patch(
            "app.services.ai.adapter.logger.exception"
        ) as logged_fallback:
            resilient = ResilientAdapter(QwenAdapter(), MockAdapter())
            result = resilient.judge_merge(
                {"text": "F08风机接地"},
                {"text": "F13风机接地"},
            )

        self.assertFalse(result["same_item"])
        self.assertEqual(result["reason_code"], "DIFF_EQUIPMENT")
        self.assertEqual(resilient.fallback_count, 1)
        self.assertIn("RuntimeError", resilient.last_error)
        fake_client.chat.completions.create.assert_called_once()
        logged_fallback.assert_called_once()


class V040WorkLogAITest(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(self.engine)
        self.db = sessionmaker(bind=self.engine)()
        self.station = Station(
            code="TARGET",
            name="目标风电场",
            aliases_json='["目标"]',
        )
        self.db.add(self.station)
        self.db.flush()
        self.batch = HandoverBatch(
            start_date="2026-08-14",
            end_date="2026-08-23",
            handover_date="2026-08-23",
        )
        self.db.add(self.batch)
        self.db.flush()
        self.meta = HandoverStationMeta(
            batch_id=self.batch.id,
            station_id=self.station.id,
        )
        self.db.add(self.meta)
        self.db.commit()

    def tearDown(self) -> None:
        self.db.close()
        self.engine.dispose()

    @staticmethod
    def _work_log(path: Path, *, completed: bool = False) -> None:
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"
        sheet.append(["实际工作日志"])
        sheet.append([
            "date", "station", "工作内容", "完成期限", "工作要求",
            "人员安排", "是否完成", "是否列入本日计划", "工作进度",
        ])
        sheet.append([
            "2026-08-14", "目标风电场", "目标工作一", "2026-08-20",
            "按方案执行", "甲", "已完成" if completed else "未完成", "是",
            "已完成" if completed else "处理中",
        ])
        sheet.append([
            "2026-08-15", "其他风电场", "他站工作", "2026-08-20",
            "不应发送", "乙", "未完成", "是", "处理中",
        ])
        sheet.append([
            "2026-08-30", "目标风电场", "窗口外工作", "2026-09-01",
            "不应发送", "甲", "未完成", "是", "处理中",
        ])
        sheet.append([
            "2026-08-16", "目标风电场", "日定期工作检查", "2026-08-16",
            "不应发送", "甲", "已完成", "是", "已完成",
        ])
        sheet.append([
            "2026-08-17", "目标风电场", "目标工作二", "2026-08-23",
            "复核设备", "丙", "未完成", "是", "待复核",
        ])
        book.save(path)

    def test_work_log_sends_only_station_window_nonperiodic_rows_to_ai(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "filtered.xlsx"
            self._work_log(workbook)
            captured: dict = {}

            class CapturingAI:
                def enrich_preview_rows(self, rows, context):
                    captured["rows"] = json.loads(json.dumps(rows, ensure_ascii=False))
                    captured["context"] = dict(context)
                    return {
                        "items": [],
                        "usage": {"total_tokens": 8},
                        "model": "qwen3.8-flash",
                    }

            with mock.patch.multiple(
                config,
                IMPORT_DIR=root / "archive",
                AI_MODE="qwen",
                QWEN_API_KEY="fake-key",
                QWEN_MODEL="qwen3.8-flash",
            ), mock.patch.object(sections, "get_ai", return_value=CapturingAI()):
                preview = sections.create_preview(
                    self.db, self.batch.id, self.meta.id, workbook
                )

        self.assertEqual(preview["ai"]["status"], "success")
        self.assertEqual(preview["ai"]["usage"]["total_tokens"], 8)
        self.assertEqual(captured["context"]["station"], "目标风电场")
        self.assertEqual(
            {row["title_snapshot"] for row in captured["rows"]},
            {"目标工作一", "目标工作二"},
        )
        self.assertNotIn(
            "其他风电场",
            json.dumps(captured["rows"], ensure_ascii=False),
        )
        self.assertEqual(
            {row["title_snapshot"] for row in preview["rows"]},
            {"目标工作一", "目标工作二"},
        )

    def test_ai_failure_keeps_deterministic_preview_and_records_fallback(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "fallback.xlsx"
            self._work_log(workbook, completed=True)

            class FailedAI:
                def enrich_preview_rows(self, _rows, _context):
                    return {
                        "items": [],
                        "usage": {},
                        "model": "qwen3.8-flash",
                        "error": "RuntimeError: mock outage",
                    }

            with mock.patch.multiple(
                config,
                IMPORT_DIR=root / "archive",
                AI_MODE="qwen",
                QWEN_API_KEY="fake-key",
                QWEN_MODEL="qwen3.8-flash",
            ), mock.patch.object(sections, "get_ai", return_value=FailedAI()):
                preview = sections.create_preview(
                    self.db, self.batch.id, self.meta.id, workbook
                )

        self.assertEqual(preview["ai"]["status"], "fallback")
        self.assertEqual(preview["ai"]["applied"], 0)
        rows = {row["title_snapshot"]: row for row in preview["rows"]}
        self.assertEqual(rows["目标工作一"]["status"], "completed")
        self.assertEqual(rows["目标工作一"]["section"], "important")
        self.assertEqual(rows["目标工作一"]["completed_by"], "甲")
        warning_text = json.dumps(preview["warnings"], ensure_ascii=False)
        self.assertIn("自动回退到确定性规则", warning_text)


class V040LocalWorkbookAcceptanceTest(unittest.TestCase):
    """Optional acceptance against the user's local workbook, never Git data."""

    def test_local_workbook_previews_without_writing_formal_items(self):
        source = PROJECT_ROOT.parent / "导入模板.xlsx"
        if not source.exists():
            self.skipTest("仓库上一级未找到本地验收工作簿")

        book = load_workbook(source, data_only=True, read_only=False)
        if "Sheet1" not in book.sheetnames:
            self.skipTest("本地工作簿不含 Sheet1")
        sheet = book["Sheet1"]
        headers = sections._headers(sheet, 2)
        date_column = headers.get("date") or headers.get("日期") or 1
        station_column = headers.get("station") or headers.get("场站") or 2
        title_column = headers.get("工作内容")
        if title_column is None:
            self.skipTest("本地工作簿缺少工作内容列")

        current_day = None
        current_station = ""
        row_state: dict[int, tuple[object, str]] = {}
        candidates: list[tuple[str, str]] = []
        selected_station = ""
        selected_normalized = ""
        for row_no in range(3, sheet.max_row + 1):
            raw_day = sheet.cell(row_no, date_column).value
            raw_station = sheet.cell(row_no, station_column).value
            if raw_day not in (None, ""):
                current_day = raw_day
            if raw_station not in (None, ""):
                current_station = str(raw_station).strip()
            row_state[row_no] = (current_day, current_station)
            title = sheet.cell(row_no, title_column).value
            parsed_date, error = sections._iso_date(current_day, default_year=2026)
            if not title or error or not parsed_date or not current_station:
                continue
            if not selected_station:
                selected_station = current_station
                selected_normalized = sections._normalize(current_station)
            if sections._normalize(current_station) == selected_normalized:
                candidates.append((parsed_date, current_station))
        if not candidates:
            self.skipTest("本地工作簿中没有可验收的场站日期行")

        start_date = min(value[0] for value in candidates)
        end_date = max(value[0] for value in candidates)
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            engine = create_engine("sqlite:///:memory:")
            Base.metadata.create_all(engine)
            db = sessionmaker(bind=engine)()
            station = Station(
                code="LOCAL_ACCEPTANCE",
                name=selected_station,
                aliases_json="[]",
            )
            db.add(station)
            db.flush()
            batch = HandoverBatch(
                start_date=start_date,
                end_date=end_date,
                handover_date=end_date,
            )
            db.add(batch)
            db.flush()
            meta = HandoverStationMeta(batch_id=batch.id, station_id=station.id)
            db.add(meta)
            db.commit()
            with mock.patch.multiple(
                config,
                IMPORT_DIR=root / "archive",
                AI_MODE="mock",
                QWEN_API_KEY="",
            ):
                preview = sections.create_preview(db, batch.id, meta.id, source)

            self.assertEqual(preview["parser_key"], "work_log")
            self.assertGreater(len(preview["rows"]), 0)
            self.assertEqual(db.query(HandoverItem).count(), 0)
            self.assertEqual(db.query(ExternalAssessment).count(), 0)
            for row in preview["rows"]:
                row_no = int(row["source"]["row_no"])
                source_day, source_station = row_state[row_no]
                parsed_date, error = sections._iso_date(
                    source_day, default_year=int(start_date[:4])
                )
                self.assertIsNone(error, f"源行 {row_no} 日期应可解析")
                self.assertTrue(
                    start_date <= parsed_date <= end_date,
                    f"源行 {row_no} 应在所选日期窗内",
                )
                self.assertTrue(
                    sections._station_matches(station, source_station),
                    f"源行 {row_no} 应属于所选场站",
                )
            db.close()
            engine.dispose()


class V040BackupTest(unittest.TestCase):
    def test_online_sqlite_backup_is_consistent_and_manifest_is_verified(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            database = root / "live" / "handover.db"
            database.parent.mkdir(parents=True)
            live = sqlite3.connect(str(database), timeout=30)
            live.execute("PRAGMA journal_mode=WAL")
            live.execute("CREATE TABLE facts (id INTEGER PRIMARY KEY, value TEXT)")
            live.executemany(
                "INSERT INTO facts(value) VALUES (?)",
                [("第一条",), ("第二条",), ("第三条",)],
            )
            live.commit()

            snapshots = root / "snapshots"
            nas = root / "nas"
            source_engine = create_engine(
                f"sqlite:///{database.as_posix()}",
                connect_args={"timeout": 30},
            )
            try:
                with mock.patch.multiple(
                    config,
                    SNAPSHOT_DIR=snapshots,
                    NAS_BACKUP_DIR=str(nas),
                ), mock.patch.object(backup_service, "engine", source_engine):
                    result = backup_service.create_database_backup(reason="test")
            finally:
                live.close()
                source_engine.dispose()

            backup_path = Path(result["local_path"])
            manifest_path = Path(result["manifest_path"])
            self.assertTrue(backup_path.exists())
            self.assertTrue(manifest_path.exists())
            digest = hashlib.sha256(backup_path.read_bytes()).hexdigest()
            self.assertEqual(result["sha256"], digest)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["sha256"], digest)
            self.assertEqual(manifest["size"], backup_path.stat().st_size)
            self.assertEqual(manifest["nas_error"], "")

            backup = sqlite3.connect(str(backup_path))
            try:
                self.assertEqual(backup.execute("PRAGMA integrity_check").fetchone()[0], "ok")
                values = [
                    row[0]
                    for row in backup.execute("SELECT value FROM facts ORDER BY id")
                ]
            finally:
                backup.close()
            self.assertEqual(values, ["第一条", "第二条", "第三条"])

            nas_path = Path(result["nas_path"])
            self.assertTrue(nas_path.exists())
            self.assertEqual(hashlib.sha256(nas_path.read_bytes()).hexdigest(), digest)


class V040ServerMigrationTest(unittest.TestCase):
    def test_stopped_server_data_can_be_relocated_without_touching_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "旧正式数据"
            target = root / "新正式数据"
            database = source / "data" / "handover.db"
            database.parent.mkdir(parents=True)
            connection = sqlite3.connect(str(database))
            connection.execute("CREATE TABLE facts (value TEXT NOT NULL)")
            connection.executemany(
                "INSERT INTO facts(value) VALUES (?)", [("一",), ("二",)]
            )
            connection.commit()
            connection.close()
            imported = source / "imports" / "样例.xlsx"
            imported.parent.mkdir(parents=True)
            imported.write_bytes(b"sanitized-import")
            generated = source / "generated" / "样例.docx"
            generated.parent.mkdir(parents=True)
            generated.write_bytes(b"sanitized-word")
            target.mkdir()

            result = relocate_server_data(source, target)

            self.assertTrue(database.exists(), "旧目录必须保留")
            self.assertTrue((target / "data" / "handover.db").exists())
            self.assertEqual((target / "imports" / "样例.xlsx").read_bytes(), b"sanitized-import")
            self.assertEqual((target / "generated" / "样例.docx").read_bytes(), b"sanitized-word")
            migrated = sqlite3.connect(str(target / "data" / "handover.db"))
            try:
                values = [row[0] for row in migrated.execute("SELECT value FROM facts ORDER BY rowid")]
                self.assertEqual(migrated.execute("PRAGMA quick_check").fetchone()[0], "ok")
            finally:
                migrated.close()
            self.assertEqual(values, ["一", "二"])
            self.assertTrue(Path(result["manifest_path"]).exists())
            self.assertTrue(result["source_preserved"])

    def test_v030_import_backs_up_server_and_rewrites_historical_word_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            legacy = root / "old" / "JXHandover"
            target = root / "server" / "JXHandoverServer"
            legacy_db = legacy / "data" / "handover.db"
            target_db = target / "data" / "handover.db"
            legacy_db.parent.mkdir(parents=True)
            target_db.parent.mkdir(parents=True)

            relative_word = Path("TEST") / "202608" / "发布历史" / "历史_V001.docx"
            legacy_word = legacy / "generated" / relative_word
            legacy_word.parent.mkdir(parents=True)
            legacy_word.write_bytes(b"legacy-word")
            # Exercise collision handling: the existing server file is kept,
            # while the imported snapshot must point to the renamed old file.
            current_word = target / "generated" / relative_word
            current_word.parent.mkdir(parents=True)
            current_word.write_bytes(b"current-server-word")

            old = sqlite3.connect(str(legacy_db))
            old.execute(
                "CREATE TABLE document_snapshots "
                "(id TEXT PRIMARY KEY, docx_path TEXT NOT NULL)"
            )
            old.execute(
                "INSERT INTO document_snapshots VALUES (?, ?)",
                ("snap_old", str(legacy_word)),
            )
            old.commit()
            old.close()

            current = sqlite3.connect(str(target_db))
            current.execute("CREATE TABLE marker (value TEXT)")
            current.execute("INSERT INTO marker VALUES ('server-before-import')")
            current.commit()
            current.close()
            Path(f"{target_db}-wal").write_bytes(b"")

            result = migrate_v030_data(legacy, target)

            self.assertTrue(Path(result["previous_server_backup"]).exists())
            self.assertTrue(Path(result["manifest_path"]).exists())
            self.assertFalse(Path(f"{target_db}-wal").exists())
            migrated = sqlite3.connect(str(target_db))
            try:
                rewritten = Path(migrated.execute(
                    "SELECT docx_path FROM document_snapshots WHERE id='snap_old'"
                ).fetchone()[0])
            finally:
                migrated.close()
            self.assertTrue(rewritten.exists())
            self.assertEqual(rewritten.read_bytes(), b"legacy-word")
            self.assertNotEqual(rewritten, current_word)
            previous = sqlite3.connect(result["previous_server_backup"])
            try:
                self.assertEqual(
                    previous.execute("SELECT value FROM marker").fetchone()[0],
                    "server-before-import",
                )
            finally:
                previous.close()


if __name__ == "__main__":
    unittest.main()
