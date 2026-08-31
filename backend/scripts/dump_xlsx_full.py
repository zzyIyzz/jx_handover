# -*- coding: utf-8 -*-
"""全量转储传入 XLSX：所有 sheet、所有行、所有列（含合并单元格展开）。"""
import os
from pathlib import Path
import sys
from openpyxl import load_workbook

OUT = os.path.join(os.environ.get("TEMP", "."), "xlsx_full_dump.txt")

source = Path(sys.argv[1]).expanduser().resolve() if len(sys.argv) > 1 else None
if source is None or not source.is_file():
    raise SystemExit("用法：python dump_xlsx_full.py <待检查的.xlsx>")
wb = load_workbook(source, data_only=True)
lines = [f"sheets: {wb.sheetnames}"]
for ws in wb.worksheets:
    lines.append(f"\n########## SHEET: {ws.title} "
                 f"rows={ws.max_row} cols={ws.max_column} ##########")
    # 合并单元格值展开
    merged = {}
    for mr in ws.merged_cells.ranges:
        top = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged[(r, c)] = top
    for row in ws.iter_rows():
        vals = []
        for cell in row:
            v = cell.value
            if v is None:
                v = merged.get((cell.row, cell.column))
            s = str(v).strip().replace("\n", "⏎") if v is not None else ""
            vals.append(s)
        lines.append(f"R{row[0].row}: " + " | ".join(vals))
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("OK ->", OUT)
