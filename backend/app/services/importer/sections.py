"""Preview-first XLSX import for chapters three, four and five."""
from __future__ import annotations

import hashlib
import json
import re
import shutil
import uuid
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

from app import config
from app.models import (
    ExternalAssessment,
    HandoverBatch,
    HandoverItem,
    HandoverStationMeta,
    ImportJob,
    SectionImportPreview,
    SourceRecord,
    Station,
    WorkItem,
    WorkItemUpdate,
    new_id,
    now_iso,
)


STANDARD_SHEETS = {
    "第三章-重点工作": "important",
    "第四章-需交接": "handover",
    "第五章-外委考核": "external",
}
ITEM_STATUSES = {"pending", "in_progress", "blocked", "completed", "unknown"}
PRIORITIES = {"urgent", "important", "normal"}

STATUS_ALIASES = {
    "已完成": "completed",
    "完成": "completed",
    "未完成": "in_progress",
    "进行中": "in_progress",
    "处理中": "in_progress",
    "受阻": "blocked",
    "阻塞": "blocked",
    "待启动": "pending",
    "未开始": "pending",
    "待处理": "pending",
    "未知": "unknown",
    "是": "completed",
    "否": "in_progress",
    "Y": "completed",
    "N": "in_progress",
}
PRIORITY_ALIASES = {
    "紧急": "urgent",
    "重点": "important",
    "普通": "normal",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize(value: Any) -> str:
    return re.sub(r"[\s，。、；：,.:;（）()【】\[\]]+", "", _clean(value)).lower()


def _iso_date(value: Any, *, default_year: int | None = None) -> tuple[str | None, str | None]:
    if value in (None, "", "—", "-"):
        return None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), None
    if isinstance(value, date):
        return value.isoformat(), None
    text = _clean(value)
    patterns = (
        r"(?P<year>\d{4})[./年-](?P<month>\d{1,2})[./月-](?P<day>\d{1,2})",
        r"(?P<month>\d{1,2})[./月](?P<day>\d{1,2})",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        year = int(match.groupdict().get("year") or default_year or 0)
        if not year:
            return None, "日期缺少年份"
        try:
            return date(year, int(match.group("month")), int(match.group("day"))).isoformat(), None
        except ValueError:
            return None, f"无效日期：{text}"
    return None, f"无法识别日期：{text}"


def _status(value: Any) -> str:
    text = _clean(value)
    if text in ITEM_STATUSES:
        return text
    if text in STATUS_ALIASES:
        return STATUS_ALIASES[text]
    for label, status in sorted(STATUS_ALIASES.items(),
                                key=lambda item: len(item[0]), reverse=True):
        if len(label) > 1 and label in text:
            return status
    return "unknown"


def _priority(value: Any) -> str:
    text = _clean(value)
    if text in PRIORITIES:
        return text
    for label, priority in PRIORITY_ALIASES.items():
        if label in text:
            return priority
    return "normal"


def _base_row(sheet: str, row_no: int, kind: str, section: str) -> dict:
    return {
        "preview_key": uuid.uuid4().hex,
        "kind": kind,
        "section": section,
        "include": True,
        "valid": True,
        "duplicate": False,
        "errors": [],
        "warnings": [],
        "source": {"sheet": sheet, "row_no": row_no, "raw": {}},
    }


def _headers(sheet, row_no: int) -> dict[str, int]:
    return {
        _clean(sheet.cell(row_no, column).value).replace(" ", ""): column
        for column in range(1, sheet.max_column + 1)
        if _clean(sheet.cell(row_no, column).value)
    }


def _value(sheet, row_no: int, headers: dict[str, int], *names: str) -> Any:
    for name in names:
        if name in headers:
            return sheet.cell(row_no, headers[name]).value
    return None


def _parse_standard(book, batch: HandoverBatch) -> tuple[list[dict], list[dict]]:
    rows: list[dict] = []
    warnings: list[dict] = []
    for sheet_name, section in STANDARD_SHEETS.items():
        if sheet_name not in book.sheetnames:
            warnings.append({"sheet": sheet_name, "field": "工作表", "reason": "缺少工作表"})
            continue
        sheet = book[sheet_name]
        headers = _headers(sheet, 1)
        required_headers = (
            {"外委单位", "工作内容", "考核情况"}
            if section == "external"
            else {"工作内容", "完成情况", "优先级"}
        )
        for missing_header in sorted(required_headers - set(headers)):
            warnings.append({
                "sheet": sheet_name,
                "field": missing_header,
                "reason": "缺少必要列",
            })
        for row_no in range(2, sheet.max_row + 1):
            title = _clean(_value(sheet, row_no, headers, "工作内容"))
            contractor = _clean(_value(sheet, row_no, headers, "外委单位"))
            assessment = _clean(_value(sheet, row_no, headers, "考核情况"))
            if not title and not contractor and not assessment:
                continue
            if section == "external":
                row = _base_row(sheet_name, row_no, "external", "external")
                row.update({
                    "contractor": contractor,
                    "work_content": title,
                    "assessment": assessment,
                    "remark": _clean(_value(sheet, row_no, headers, "备注")),
                })
                for value, label in (
                    (contractor, "外委单位"),
                    (title, "工作内容"),
                    (row["assessment"], "考核情况"),
                ):
                    if value:
                        continue
                    row["valid"] = False
                    row["errors"].append(f"{label}不能为空")
            else:
                raw_status = _value(sheet, row_no, headers, "完成情况", "状态")
                status = _status(raw_status)
                start_date, start_error = _iso_date(
                    _value(sheet, row_no, headers, "开始时间"), default_year=int(batch.start_date[:4]))
                end_date, end_error = _iso_date(
                    _value(sheet, row_no, headers, "结束时间"), default_year=int(batch.end_date[:4]))
                row = _base_row(sheet_name, row_no, "item", section)
                row.update({
                    "title_snapshot": title,
                    "status": status,
                    "priority": _priority(_value(sheet, row_no, headers, "优先级")),
                    "completed_by": _clean(_value(sheet, row_no, headers, "完成人")),
                    "previous_owner": _clean(_value(sheet, row_no, headers, "交接前责任人")),
                    "next_owner": _clean(_value(sheet, row_no, headers, "交接后责任人")),
                    "start_date": start_date,
                    "end_date": end_date,
                    "summary": "",
                    "latest_progress": _clean(_value(sheet, row_no, headers, "备注")),
                    "blocker": "",
                    "next_action": "",
                })
                if not title:
                    row["valid"] = False
                    row["errors"].append("工作内容不能为空")
                for error in (start_error, end_error):
                    if error:
                        row["valid"] = False
                        row["errors"].append(error)
                if status == "unknown" and _clean(raw_status):
                    row["warnings"].append(f"无法确定完成情况：{_clean(raw_status)}")
                if section == "important" and not row["completed_by"]:
                    row["valid"] = False
                    row["errors"].append("第三章事项缺少完成人")
            row["source"]["raw"] = {
                header: sheet.cell(row_no, column).value
                for header, column in headers.items()
            }
            rows.append(row)
    return rows, warnings


def _station_matches(station: Station, text: str) -> bool:
    aliases = json.loads(station.aliases_json or "[]") + [station.name]
    normalized = _normalize(text)
    return any(_normalize(alias) in normalized for alias in aliases if alias)


def _style_priority(cell) -> tuple[str, str | None]:
    font_rgb = ""
    if cell.font.color is not None and cell.font.color.type == "rgb":
        font_rgb = str(cell.font.color.rgb or "").upper()
    if font_rgb in {"FFFF0000", "FFDE3C36", "00FF0000"}:
        return "urgent", "依据工作内容红字标记识别为紧急，请在预览中确认"
    fill_rgb = ""
    if cell.fill.fgColor.type == "rgb":
        fill_rgb = str(cell.fill.fgColor.rgb or "").upper()
    colored_fill = cell.fill.patternType == "solid" and fill_rgb not in {
        "", "00000000", "FFFFFFFF", "00FFFFFF"
    }
    if colored_fill or bool(cell.font.bold):
        return "important", "依据工作内容底色/加粗标记识别为重点，请在预览中确认"
    return "normal", None


def _parse_work_log(book, batch: HandoverBatch, station: Station) -> tuple[list[dict], list[dict]]:
    sheet = book["Sheet1"]
    headers = _headers(sheet, 2)
    required = {"工作内容", "人员安排", "是否完成", "工作进度"}
    missing = sorted(required - set(headers))
    if missing:
        raise HTTPException(422, {
            "code": "WORK_LOG_COLUMNS_MISSING",
            "message": "工作日志缺少必要列。",
            "columns": missing,
        })

    current_day: Any = None
    current_station = ""
    # The real workbook uses columns A/B for merged date/station values but
    # labels A2 as the monthly-plan title instead of providing two headers.
    date_column = headers.get("date") or headers.get("日期") or 1
    station_column = headers.get("station") or headers.get("场站") or 2
    latest: OrderedDict[str, dict] = OrderedDict()
    first_seen: dict[str, str | None] = {}
    global_warnings: list[dict] = []
    for row_no in range(3, sheet.max_row + 1):
        if sheet.cell(row_no, date_column).value not in (None, ""):
            current_day = sheet.cell(row_no, date_column).value
        if sheet.cell(row_no, station_column).value not in (None, ""):
            current_station = _clean(sheet.cell(row_no, station_column).value)
        title = _clean(_value(sheet, row_no, headers, "工作内容"))
        if not title or not _station_matches(station, current_station):
            continue
        plan_date, date_error = _iso_date(current_day, default_year=int(batch.start_date[:4]))
        if date_error or not plan_date or not (batch.start_date <= plan_date <= batch.end_date):
            continue
        # Periodic work belongs to chapter six and must never be silently copied
        # into chapter three/four.
        if "定期工作" in title and any(word in title for word in ("日", "周", "月", "季", "年")):
            continue

        key = _normalize(title)
        raw_status = _value(sheet, row_no, headers, "是否完成")
        status = _status(raw_status)
        section = "important" if status == "completed" else "handover"
        due_date, due_error = _iso_date(
            _value(sheet, row_no, headers, "完成期限"),
            default_year=int(batch.end_date[:4]),
        )
        priority, priority_warning = _style_priority(
            sheet.cell(row_no, headers["工作内容"]))
        text_priority = _priority(title)
        if text_priority != "normal":
            priority = text_priority
            priority_warning = None
        personnel = _clean(_value(sheet, row_no, headers, "人员安排"))
        progress = _clean(_value(sheet, row_no, headers, "工作进度"))
        requirement = _clean(_value(sheet, row_no, headers, "工作要求"))
        row = _base_row(sheet.title, row_no, "item", section)
        row.update({
            "title_snapshot": title,
            "status": status,
            "priority": priority,
            "completed_by": personnel if status == "completed" else "",
            "previous_owner": personnel if status != "completed" else "",
            "next_owner": "",
            "start_date": first_seen.get(key) or plan_date,
            "end_date": plan_date if status == "completed" else due_date,
            "summary": requirement,
            "latest_progress": progress,
            "blocker": "",
            "next_action": "",
            "planned_today": _clean(_value(sheet, row_no, headers, "是否列入本日计划")),
        })
        first_seen.setdefault(key, plan_date)
        row["start_date"] = first_seen[key]
        if due_error:
            row["warnings"].append(due_error)
            row["end_date"] = None
        if status == "unknown":
            row["warnings"].append("完成情况为空或无法识别，请人工选择")
        if priority_warning:
            row["warnings"].append(priority_warning)
        if "," in personnel or "，" in personnel:
            row["warnings"].append("存在多名人员，请确认第三章完成人或第四章责任人")
        row["source"]["raw"] = {
            header: sheet.cell(row_no, column).value
            for header, column in headers.items()
        }
        latest[key] = row

    rows = list(latest.values())
    if not rows:
        global_warnings.append({
            "sheet": sheet.title,
            "field": "场站/日期",
            "reason": f"未找到 {station.name} 在 {batch.start_date} 至 {batch.end_date} 的工作记录",
        })
    else:
        global_warnings.append({
            "sheet": sheet.title,
            "field": "去重规则",
            "reason": "同一事项在班次内按工作内容合并，保留最后一个工作日的状态和进度",
        })
    return rows, global_warnings


def _detect_adapter(book) -> str:
    if set(STANDARD_SHEETS).issubset(set(book.sheetnames)):
        return "standard_template"
    if "Sheet1" in book.sheetnames:
        headers = _headers(book["Sheet1"], 2)
        if {"工作内容", "是否完成", "工作进度"}.issubset(headers):
            return "work_log"
    raise HTTPException(422, {
        "code": "UNSUPPORTED_XLSX_LAYOUT",
        "message": "无法识别这份 XLSX。请使用标准导入模板或实际工作日志格式。",
    })


def _mark_duplicates(db, meta_id: str, rows: list[dict]) -> None:
    existing_items = (db.query(HandoverItem)
                      .filter(HandoverItem.station_meta_id == meta_id).all())
    item_keys = {
        (_normalize(item.title_snapshot), item.section,
         item.start_date or "", item.end_date or "")
        for item in existing_items
    }
    existing_external = (db.query(ExternalAssessment)
                         .filter(ExternalAssessment.station_meta_id == meta_id).all())
    external_keys = {
        (_normalize(row.contractor), _normalize(row.work_content),
         _normalize(row.assessment), _normalize(row.remark))
        for row in existing_external
    }
    seen: set[tuple] = set()
    for row in rows:
        row["duplicate"] = False
        if row["kind"] == "item":
            key = (_normalize(row.get("title_snapshot")), row.get("section"),
                   row.get("start_date") or "", row.get("end_date") or "")
        else:
            key = (_normalize(row.get("contractor")),
                   _normalize(row.get("work_content")),
                   _normalize(row.get("assessment")),
                   _normalize(row.get("remark")))
        duplicate = key in seen or key in (item_keys if row["kind"] == "item" else external_keys)
        seen.add(key)
        if duplicate:
            row["duplicate"] = True
            row["include"] = False
            row["warnings"].append("检测到重复记录，默认跳过")


def create_preview(db, batch_id: str, meta_id: str, source_path: Path) -> dict:
    batch = db.get(HandoverBatch, batch_id)
    meta = db.get(HandoverStationMeta, meta_id)
    if batch is None or meta is None or meta.batch_id != batch_id:
        raise HTTPException(404, "交接班或场站信息不存在")
    station = db.get(Station, meta.station_id)
    if station is None:
        raise HTTPException(404, "场站不存在")

    source_hash = _sha256(source_path)
    archive_dir = config.IMPORT_DIR / batch_id
    archive_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^\w.()（）-]+", "_", source_path.name, flags=re.UNICODE)
    archived = archive_dir / f"{source_hash[:12]}_{safe_name}"
    if not archived.exists():
        shutil.copy2(source_path, archived)

    book = load_workbook(archived, data_only=True, read_only=False)
    parser_key = _detect_adapter(book)
    if parser_key == "standard_template":
        rows, warnings = _parse_standard(book, batch)
    else:
        rows, warnings = _parse_work_log(book, batch, station)
    _mark_duplicates(db, meta_id, rows)

    job = ImportJob(
        source_type="section_xlsx",
        file_name=source_path.name,
        file_sha256=source_hash,
        stored_path=str(archived),
        parser_key=parser_key,
        status="success",
        row_count=len(rows),
        finished_at=now_iso(),
    )
    db.add(job)
    db.flush()
    preview = SectionImportPreview(
        batch_id=batch_id,
        station_meta_id=meta_id,
        import_job_id=job.id,
        parser_key=parser_key,
        source_file_name=source_path.name,
        source_sha256=source_hash,
        normalized_json=json.dumps(rows, ensure_ascii=False, default=str),
        warnings_json=json.dumps(warnings, ensure_ascii=False, default=str),
    )
    db.add(preview)
    db.commit()
    return _preview_response(preview)


def _preview_response(preview: SectionImportPreview) -> dict:
    rows = json.loads(preview.normalized_json or "[]")
    return {
        "id": preview.id,
        "batch_id": preview.batch_id,
        "station_meta_id": preview.station_meta_id,
        "parser_key": preview.parser_key,
        "source_file_name": preview.source_file_name,
        "source_sha256": preview.source_sha256,
        "status": preview.status,
        "rows": rows,
        "warnings": json.loads(preview.warnings_json or "[]"),
        "summary": {
            "total": len(rows),
            "important": sum(r.get("section") == "important" for r in rows),
            "handover": sum(r.get("section") == "handover" for r in rows),
            "external": sum(r.get("kind") == "external" for r in rows),
            "invalid": sum(not r.get("valid", True) for r in rows),
            "duplicate": sum(r.get("duplicate", False) for r in rows),
        },
        "result": json.loads(preview.result_json or "{}"),
    }


def get_preview(db, preview_id: str) -> dict:
    preview = db.get(SectionImportPreview, preview_id)
    if preview is None:
        raise HTTPException(404, "导入预览不存在")
    return _preview_response(preview)


def _validate_commit_row(row: dict, batch: HandoverBatch) -> list[str]:
    errors: list[str] = []
    if row.get("kind") == "item":
        if not _clean(row.get("title_snapshot")):
            errors.append("工作内容不能为空")
        if row.get("section") not in {"important", "handover"}:
            errors.append("事项章节无效")
        if row.get("status") not in ITEM_STATUSES:
            errors.append("事项状态无效")
        if row.get("priority") not in PRIORITIES:
            errors.append("事项优先级无效")
        if row.get("section") == "important" and not _clean(row.get("completed_by")):
            errors.append("第三章事项缺少完成人")
        for field, label, year in (
            ("start_date", "开始时间", int(batch.start_date[:4])),
            ("end_date", "结束时间", int(batch.end_date[:4])),
        ):
            value = row.get(field)
            if value in (None, ""):
                continue
            parsed, error = _iso_date(value, default_year=year)
            if error:
                errors.append(f"{label}{error}")
            else:
                row[field] = parsed
    elif row.get("kind") == "external":
        for field, label in (
            ("contractor", "外委单位"),
            ("work_content", "工作内容"),
            ("assessment", "考核情况"),
        ):
            if not _clean(row.get(field)):
                errors.append(f"外委考核{label}不能为空")
    else:
        errors.append("导入记录类型无效")
    return errors


def commit_preview(db, batch_id: str, preview_id: str,
                   edited_rows: list[dict] | None = None) -> dict:
    preview = db.get(SectionImportPreview, preview_id)
    if preview is None or preview.batch_id != batch_id:
        raise HTTPException(404, "导入预览不存在")
    if preview.status == "committed":
        return json.loads(preview.result_json or "{}")
    stored_rows = json.loads(preview.normalized_json or "[]")
    if edited_rows is None:
        rows = stored_rows
    else:
        originals = {row.get("preview_key"): row for row in stored_rows}
        rows = []
        for edited in edited_rows:
            original = originals.get(edited.get("preview_key"))
            if original is None:
                continue
            row = dict(edited)
            row["preview_key"] = original.get("preview_key")
            row["kind"] = original.get("kind")
            row["source"] = original.get("source")
            rows.append(row)
    meta = db.get(HandoverStationMeta, preview.station_meta_id)
    batch = db.get(HandoverBatch, batch_id)
    if meta is None or batch is None:
        raise HTTPException(404, "交接班或场站信息不存在")

    for row in rows:
        errors = _validate_commit_row(row, batch)
        row["errors"] = errors
        row["valid"] = not errors
    _mark_duplicates(db, meta.id, rows)

    created_items = 0
    created_external = 0
    skipped = 0
    max_orders = {
        "important": (db.query(HandoverItem.sort_order)
                      .filter(HandoverItem.station_meta_id == meta.id,
                              HandoverItem.section == "important")
                      .order_by(HandoverItem.sort_order.desc()).first()),
        "handover": (db.query(HandoverItem.sort_order)
                     .filter(HandoverItem.station_meta_id == meta.id,
                             HandoverItem.section == "handover")
                     .order_by(HandoverItem.sort_order.desc()).first()),
        "external": (db.query(ExternalAssessment.sort_order)
                     .filter(ExternalAssessment.station_meta_id == meta.id)
                     .order_by(ExternalAssessment.sort_order.desc()).first()),
    }
    order_values = {key: (value[0] if value else 0) for key, value in max_orders.items()}

    for row in rows:
        if not row.get("include", True) or row.get("duplicate", False) or not row.get("valid", True):
            skipped += 1
            continue
        source = row.get("source") or {}
        raw_json = json.dumps(source.get("raw") or {}, ensure_ascii=False, default=str)
        if row["kind"] == "external":
            order_values["external"] += 10
            db.add(ExternalAssessment(
                batch_id=batch_id,
                station_meta_id=meta.id,
                contractor=_clean(row.get("contractor")),
                work_content=_clean(row.get("work_content")),
                assessment=_clean(row.get("assessment")),
                remark=_clean(row.get("remark")),
                sort_order=order_values["external"],
                source_type="xlsx",
                source_json=json.dumps({
                    "import_job_id": preview.import_job_id,
                    "source_sha256": preview.source_sha256,
                    "sheet": source.get("sheet"),
                    "row_no": source.get("row_no"),
                    "raw": source.get("raw") or {},
                }, ensure_ascii=False, default=str),
            ))
            created_external += 1
            continue

        title = _clean(row.get("title_snapshot"))
        start_date = row.get("start_date") or batch.start_date
        status = row.get("status")
        priority = row.get("priority")
        section = row.get("section")
        normalized = _normalize(title + _clean(row.get("latest_progress")))
        content_hash = hashlib.sha256(
            f"{preview.source_sha256}|{source.get('sheet')}|{source.get('row_no')}|{normalized}".encode("utf-8")
        ).hexdigest()
        source_record = SourceRecord(
            import_job_id=preview.import_job_id,
            source_type="section_xlsx",
            source_date=start_date,
            station_id=meta.station_id,
            sheet_name=_clean(source.get("sheet")),
            row_no=source.get("row_no"),
            raw_text=raw_json,
            normalized_text=normalized,
            raw_json=raw_json,
            content_hash=content_hash,
        )
        db.add(source_record)
        db.flush()
        work = WorkItem(
            station_id=meta.station_id,
            canonical_title=title,
            canonical_key="",
            status=status,
            priority=priority,
            first_seen_date=start_date,
            last_seen_date=row.get("end_date") or batch.handover_date,
            is_closed=1 if status == "completed" else 0,
        )
        db.add(work)
        db.flush()
        progress = _clean(row.get("latest_progress"))
        db.add(WorkItemUpdate(
            work_item_id=work.id,
            source_record_id=source_record.id,
            update_date=row.get("end_date") or batch.handover_date,
            action_text=_clean(row.get("summary")),
            progress_text=progress,
            status_hint=status,
        ))
        order_values[section] += 10
        db.add(HandoverItem(
            batch_id=batch_id,
            station_meta_id=meta.id,
            work_item_id=work.id,
            title_snapshot=title,
            status=status,
            priority=priority,
            section=section,
            completed_by=_clean(row.get("completed_by")),
            sort_order=order_values[section],
            summary=_clean(row.get("summary")),
            latest_progress=progress,
            blocker=_clean(row.get("blocker")),
            next_action=_clean(row.get("next_action")),
            previous_owner=_clean(row.get("previous_owner")),
            next_owner=_clean(row.get("next_owner")),
            start_date=row.get("start_date"),
            end_date=row.get("end_date"),
            source_ids_json=json.dumps([source_record.id]),
            review_status="approved",
            human_edited=1,
        ))
        created_items += 1

    result = {
        "preview_id": preview.id,
        "created_items": created_items,
        "created_external_assessments": created_external,
        "skipped": skipped,
        "committed": True,
    }
    preview.normalized_json = json.dumps(rows, ensure_ascii=False, default=str)
    preview.status = "committed"
    preview.result_json = json.dumps(result, ensure_ascii=False)
    preview.committed_at = now_iso()
    db.commit()
    return result
