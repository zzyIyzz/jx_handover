"""腾讯文档导出 XLSX 导入：日期清洗、场站识别、去重入库。

清洗原则：
- 日期统一为 ISO 8601；无法确定年份时标记 DATE_UNRESOLVED 交人工，不让 AI 猜。
- 场站识别优先级：显式场站列 -> Sheet 名 -> 别名词典 -> 正文关键词 -> 无法识别。
- 内容哈希去重，同一文件重复导入不产生重复记录。
"""
from __future__ import annotations

import hashlib
import json
import re
import shutil
from datetime import date
import math
from pathlib import Path

from openpyxl import load_workbook

from app import config
from app.models import ImportJob, SourceRecord, Station, new_id, now_iso

_FULL_DATE = re.compile(r"(\d{4})\s*[./年\-]\s*(\d{1,2})\s*[./月\-]\s*(\d{1,2})")
_MD_CN = re.compile(r"(\d{1,2})\s*月\s*(\d{1,2})\s*[日号]?")
_MD_DOT = re.compile(r"^(\d{1,2})[.．](\d{1,2})$")


def parse_date(text, default_year: int | None) -> tuple[str | None, bool]:
    """返回 (ISO日期或None, 是否未解析)。"""
    if text is None:
        return None, False
    if hasattr(text, "strftime"):
        try:
            return text.strftime("%Y-%m-%d"), False
        except Exception:
            pass
    s = str(text).strip()
    if not s:
        return None, False

    m = _FULL_DATE.search(s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = _MD_CN.search(s) or _MD_DOT.match(s)
        if not m:
            # 有日期样式的文字但解析失败 -> 人工处理
            if re.search(r"\d", s) and re.search(r"[日月./]", s):
                return None, True
            return None, False
        if default_year is None:
            return None, True  # DATE_UNRESOLVED
        y = default_year
        mo, d = int(m.group(1)), int(m.group(2))
    try:
        parsed = date(y, mo, d)
        return parsed.isoformat(), False
    except (TypeError, ValueError):
        return None, True


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", "", str(text)).strip()


def _station_alias_map(db) -> dict[int, list[str]]:
    result: dict[int, list[str]] = {}
    for st in db.query(Station).filter(Station.is_active == 1).all():
        aliases = json.loads(st.aliases_json or "[]")
        result[st.id] = [a.lower() for a in aliases] + [st.name.lower()]
    return result


def identify_station(db, *, station_col: str | None, sheet_name: str | None,
                     text: str) -> int | None:
    alias_map = _station_alias_map(db)

    def match(s: str | None) -> int | None:
        if not s:
            return None
        s_low = str(s).lower()
        for sid, aliases in alias_map.items():
            if any(a and a in s_low for a in aliases):
                return sid
        return None

    # 优先级：显式场站列 -> Sheet名 -> 正文关键词
    sid = match(station_col)
    if sid:
        return sid
    sid = match(sheet_name)
    if sid:
        return sid
    return match(text)


def _content_hash(source_date: str | None, station_id: int | None,
                  normalized: str) -> str:
    raw = f"{source_date or ''}|{station_id or ''}|{normalized}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _find_col(columns: list[str], keywords: list[str]) -> str | None:
    for col in columns:
        c = str(col)
        if any(k in c for k in keywords):
            return col
    return None


def _is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and not value.strip()


def _sheet_rows(sheet):
    """Return stable header names and a streaming iterator of row dictionaries."""
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator, None)
    if header is None:
        return [], iter(())
    columns: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(header, start=1):
        base = str(value).strip() if not _is_blank(value) else f"未命名列{index}"
        count = used.get(base, 0)
        used[base] = count + 1
        columns.append(base if count == 0 else f"{base}_{count + 1}")

    def rows():
        for row_no, values in enumerate(iterator, start=2):
            padded = tuple(values) + (None,) * max(0, len(columns) - len(values))
            row = dict(zip(columns, padded[:len(columns)]))
            if any(not _is_blank(value) for value in row.values()):
                yield row_no, row

    return columns, rows()


def import_meeting_xlsx(db, file_path: Path, default_year: int | None = None,
                        force_station_code: str | None = None) -> dict:
    """导入班会记录 XLSX。"""
    job = ImportJob(
        id=new_id("imp"), source_type="tencent_xlsx", file_name=file_path.name
    )
    db.add(job)

    # 存档原文件并计算 SHA256
    data = file_path.read_bytes()
    file_sha = hashlib.sha256(data).hexdigest()
    job.file_sha256 = file_sha
    archive = config.IMPORT_DIR / f"{job.id}_{file_path.name}"
    shutil.copyfile(file_path, archive)

    force_station_id = None
    if force_station_code:
        st = db.query(Station).filter(Station.code == force_station_code).first()
        force_station_id = st.id if st else None

    inserted, skipped, date_unresolved = 0, 0, []
    try:
        book = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        job.status = "failed"
        job.error_message = str(exc)
        job.finished_at = now_iso()
        db.commit()
        return {"status": "failed", "error": str(exc)}

    try:
        for sheet in book.worksheets:
            sheet_name = str(sheet.title)
            columns, rows = _sheet_rows(sheet)
            if not columns:
                continue
            date_col = _find_col(columns, ["日期", "时间"])
            content_col = _find_col(columns, ["内容", "记录", "事项", "工作"])
            station_col = _find_col(columns, ["场站", "电站", "单位"])
            # 兜底：第一列当日期，第二列当内容
            if date_col is None and len(columns) >= 1:
                date_col = columns[0]
            if content_col is None and len(columns) >= 2:
                content_col = columns[1]

            for row_no, row in rows:
                raw_content = row.get(content_col) if content_col else None
                if _is_blank(raw_content):
                    continue
                raw_text = str(raw_content).strip()
                date_raw = row.get(date_col) if date_col else None
                source_date, is_unresolved = parse_date(date_raw, default_year)
                if is_unresolved:
                    date_unresolved.append({"sheet": sheet_name, "row": row_no,
                                            "date": str(date_raw)})

                station_val = row.get(station_col) if station_col else None
                station_id = (force_station_id
                              or identify_station(db, station_col=station_val,
                                                  sheet_name=sheet_name, text=raw_text))

                normalized = normalize_text(raw_text)
                chash = _content_hash(source_date, station_id, normalized)
                exists = (db.query(SourceRecord.id)
                          .filter(SourceRecord.content_hash == chash).first())
                if exists:
                    skipped += 1
                    continue

                db.add(SourceRecord(
                    import_job_id=job.id,
                    source_type="tencent_xlsx",
                    source_date=source_date,
                    station_id=station_id,
                    sheet_name=sheet_name,
                    row_no=row_no,
                    raw_text=raw_text,
                    normalized_text=normalized,
                    raw_json=json.dumps(
                        {str(key): (None if _is_blank(value) else str(value))
                         for key, value in row.items()}, ensure_ascii=False),
                    content_hash=chash,
                ))
                inserted += 1
    finally:
        book.close()

    job.status = "success"
    job.row_count = inserted
    job.finished_at = now_iso()
    db.commit()
    return {"status": "success", "job_id": job.id, "inserted": inserted,
            "skipped_duplicate": skipped,
            "date_unresolved": date_unresolved}


def import_monthly_plan_xlsx(db, file_path: Path, plan_month: str,
                             category: str = "monthly",
                             station_code: str | None = None,
                             default_year: int | None = None) -> dict:
    """导入月度/季度定期工作计划 XLSX。"""
    job = ImportJob(
        id=new_id("imp"), source_type="monthly_plan_xlsx", file_name=file_path.name
    )
    db.add(job)
    data = file_path.read_bytes()
    job.file_sha256 = hashlib.sha256(data).hexdigest()
    archive = config.IMPORT_DIR / f"{job.id}_{file_path.name}"
    shutil.copyfile(file_path, archive)

    station_id = None
    if station_code:
        st = db.query(Station).filter(Station.code == station_code).first()
        station_id = st.id if st else None

    inserted, skipped, date_unresolved = 0, 0, []
    try:
        book = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        job.status = "failed"
        job.error_message = str(exc)
        job.finished_at = now_iso()
        db.commit()
        return {"status": "failed", "error": str(exc)}

    from app.models import MonthlyPlanItem

    # 同一个计划文件重复导入应保持幂等，避免使用者误点两次后出现重复行。
    existing_keys = {
        (p.station_id, p.category, normalize_text(p.title),
         p.plan_start or "", p.plan_end or "")
        for p in db.query(MonthlyPlanItem).filter(
            MonthlyPlanItem.plan_month == plan_month).all()
    }

    try:
        for sheet in book.worksheets:
            sheet_name = str(sheet.title)
            columns, rows = _sheet_rows(sheet)
            if not columns:
                continue
            cat = "quarterly" if "季度" in sheet_name else category
            title_col = _find_col(columns, ["工作内容", "内容", "事项", "工作"])
            start_col = _find_col(columns, ["开始"])
            end_col = _find_col(columns, ["结束", "截止", "完成时间"])
            owner_col = _find_col(columns, ["完成人", "责任人", "负责人"])
            status_col = _find_col(columns, ["完成情况", "状态"])
            note_col = _find_col(columns, ["备注"])
            if title_col is None:
                title_col = columns[0]

            for row_index, row in rows:
                title = row.get(title_col)
                if _is_blank(title):
                    continue
                start_raw = row.get(start_col) if start_col else None
                end_raw = row.get(end_col) if end_col else None
                plan_start, start_unresolved = parse_date(start_raw, default_year)
                plan_end, end_unresolved = parse_date(end_raw, default_year)
                if start_unresolved or end_unresolved:
                    date_unresolved.append({
                        "sheet": sheet_name,
                        "row": row_index,
                        "date": str(start_raw if start_unresolved else end_raw),
                    })
                status_value = row.get(status_col) if status_col else None
                status_raw = "" if _is_blank(status_value) else str(status_value).strip()
                status = "completed" if "已完成" in status_raw or status_raw == "完成" else "pending"
                title_text = str(title).strip()
                item_key = (station_id, cat, normalize_text(title_text),
                            plan_start or "", plan_end or "")
                if item_key in existing_keys:
                    skipped += 1
                    continue
                owner_value = row.get(owner_col) if owner_col else None
                note_value = row.get(note_col) if note_col else None
                db.add(MonthlyPlanItem(
                    plan_month=plan_month,
                    scope_type="station" if station_id else "region",
                    station_id=station_id,
                    title=title_text,
                    category=cat,
                    plan_start=plan_start,
                    plan_end=plan_end,
                    owner="" if _is_blank(owner_value) else str(owner_value).strip(),
                    status=status,
                    notes="" if _is_blank(note_value) else str(note_value).strip(),
                ))
                existing_keys.add(item_key)
                inserted += 1
    finally:
        book.close()

    job.status = "success"
    job.row_count = inserted
    job.finished_at = now_iso()
    db.commit()
    return {"status": "success", "job_id": job.id, "inserted": inserted,
            "skipped_duplicate": skipped,
            "date_unresolved": date_unresolved}
